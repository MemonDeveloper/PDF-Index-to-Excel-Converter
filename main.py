import google.generativeai as genai
import PyPDF2
import openpyxl
import os

# ✅ Configure Gemini API
genai.configure(api_key="YOUR_API")
model = genai.GenerativeModel("models/gemini-2.0-flash")

# ✅ Load PDF
def extract_text_from_pdf(pdf_path):
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        texts = [page.extract_text() for page in reader.pages]
    return texts

# ✅ Process each page with Gemini
def correct_with_gemini(text):
    prompt = f"""You are a professional data parser. Extract terms and their corresponding page numbers from the following PDF index text.

Format the output as:
Term<TAB>Page Numbers
(one term per line)

Example:
activism\t64, 180–182, 184, 193
admission of wrongdoing\t357–358, 360–361

Now process this text:
{text}
"""
    response = model.generate_content(prompt)
    return response.text.strip()


# ✅ Save to Excel
def save_to_excel(data_by_page, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corrected Data"

    row = 1
    for i, page_data in enumerate(data_by_page, start=1):
        for line in page_data.splitlines():
            if "\t" in line:
                term, pages = line.split("\t", 1)
                ws.cell(row=row, column=1, value=term.strip())
                ws.cell(row=row, column=2, value=pages.strip())
                row += 1

    wb.save(output_path)

# ✅ Main Process
def process_pdf(pdf_path, output_excel):
    pages_text = extract_text_from_pdf(pdf_path)
    corrected_data = []

    for i, page in enumerate(pages_text, start=1):
        print(f"Processing Page {i}...")
        corrected = correct_with_gemini(page)
        corrected_data.append(corrected)

    save_to_excel(corrected_data, output_excel)
    print(f"\n✅ All done! Output saved to: {output_excel}")

# ▶️ Run
pdf_path = "index.pdf"
output_excel = os.path.splitext(pdf_path)[0] + ".xlsx"
process_pdf(pdf_path, output_excel)
